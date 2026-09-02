from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.exporter import export_comparison_to_excel
from core.pipeline import process_monitoring_batch
from core.vision_extractor import PriceTagExtractor


def test_full_pipeline_is_deterministic_and_in_memory(catalog_df: pd.DataFrame) -> None:
    images = [
        {"data": b"mock-1", "filename": "tag_moloko.jpg"},
        {"data": b"mock-2", "filename": "tag_maslo.jpg"},
        {"data": b"mock-3", "filename": "tag_syr.jpg"},
    ]
    results = process_monitoring_batch(
        catalog_df,
        images,
        PriceTagExtractor(provider="mock"),
        max_workers=2,
        match_threshold=65,
    )
    assert len(results) == 3
    assert all(item["extraction_status"] == "ok" for item in results)
    assert any(item["matched_sku"] for item in results)

    excel = export_comparison_to_excel(results, competitor_name="Реми")
    workbook = load_workbook(io.BytesIO(excel))
    assert workbook.sheetnames == ["Мониторинг цен", "Сводка"]
    assert workbook["Мониторинг цен"].max_row == 7


class _ExtractorWithConfidence:
    def __init__(self, confidence: object) -> None:
        self.confidence = confidence

    def extract_batch(self, images, **_kwargs):
        return [
            {
                "filename": images[0]["filename"],
                "product_name": "Молоко Простоквашино 3.2% 930мл",
                "brand": "Простоквашино",
                "weight_volume": "930мл",
                "regular_price": 100,
                "promo_price": None,
                "unit": "шт",
                "confidence": self.confidence,
                "extraction_status": "ok",
            }
        ]


@pytest.mark.parametrize("confidence", [float("nan"), float("inf"), -0.1, 1.1, "abc", None])
def test_invalid_confidence_fails_closed(catalog_df: pd.DataFrame, confidence: object) -> None:
    results = process_monitoring_batch(
        catalog_df,
        [{"data": b"image", "filename": "tag.jpg"}],
        _ExtractorWithConfidence(confidence),
    )

    assert results[0]["extraction_status"] == "error"
    assert results[0]["matched_sku"] is None
    assert results[0]["confidence"] is None
    assert "уверенность" in results[0]["match_reason"].casefold()
