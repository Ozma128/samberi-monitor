from __future__ import annotations

import io
import math

from openpyxl import load_workbook

from core.exporter import export_comparison_to_excel


def test_export_is_safe_numeric_and_has_exact_layout() -> None:
    payload = [
        {
            "matched_sku": '=HYPERLINK("https://invalid")',
            "matched_name": "+опасное имя",
            "product_name": "=1+1",
            "our_purchase_price": 80,
            "our_sale_price": 100,
            "our_promo_price": None,
            "comp_regular_price": 120,
            "comp_promo_price": None,
            "effective_diff_rub": 20,
            "price_index_effective": 120.0,
            "price_index_effective_raw": 120.0,
            "our_effective_price": 100,
            "comp_effective_price": 120,
            "status": "✅ Самбери дешевле",
            "alert": None,
            "is_dumping": False,
            "extraction_status": "ok",
        }
    ]
    workbook = load_workbook(
        io.BytesIO(export_comparison_to_excel(payload, category_name="Молочная")),
        data_only=False,
    )
    sheet = workbook["Мониторинг цен"]
    assert sheet.max_column == 12
    assert sheet["A5"].data_type == "s"
    assert sheet["A5"].value.startswith("'")
    assert sheet["B5"].value.startswith("'")
    assert sheet["C5"].data_type == "s"
    assert sheet["J5"].value == 1.2
    assert sheet["J5"].number_format == "0.0%"
    assert sheet.freeze_panes == "A5"
    assert "Молочная" in workbook["Сводка"]["A2"].value


def test_export_rejects_special_floats_as_cells() -> None:
    payload = [
        {
            "product_name": "Товар",
            "our_sale_price": math.inf,
            "comp_regular_price": math.nan,
            "price_index_effective": math.inf,
            "status": "Не определен",
        }
    ]
    workbook = load_workbook(io.BytesIO(export_comparison_to_excel(payload)))
    sheet = workbook["Мониторинг цен"]
    assert sheet["E5"].value is None
    assert sheet["G5"].value is None
    assert sheet["J5"].value is None
