"""Безопасный форматированный экспорт мониторинга в Excel."""

from __future__ import annotations

import io
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .analytics import summarize_price_index

FORMULA_PREFIXES = ("=", "+", "-", "@")
CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def sanitize_excel_text(value: Any, limit: int = 5_000) -> str:
    """Нейтрализовать формулы и управляющие символы во внешнем тексте."""

    if value is None:
        return ""
    text = CONTROL_CHARS.sub("", str(value))[:limit]
    if text.lstrip().startswith(FORMULA_PREFIXES):
        text = "'" + text
    return text


def _safe_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _first_present(item: dict[str, Any], primary: str, fallback: str) -> Any:
    value = item.get(primary)
    return item.get(fallback) if value is None else value


def export_comparison_to_excel(
    items: list[dict[str, Any]],
    competitor_name: str = "Конкурент",
    category_name: str = "Все категории",
    output_path: str | None = None,
) -> bytes:
    """Создать книгу с 12 стандартными колонками и отдельной сводкой."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Мониторинг цен"
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=15, bold=True, color="1E3A8A")
    meta_font = Font(name="Calibri", size=10, italic=True, color="555555")
    green_fill = PatternFill("solid", fgColor="E6F4EA")
    green_font = Font(name="Calibri", size=10, bold=True, color="137333")
    red_fill = PatternFill("solid", fgColor="FCE8E6")
    red_font = Font(name="Calibri", size=10, bold=True, color="C5221F")
    blue_fill = PatternFill("solid", fgColor="E8F0FE")
    blue_font = Font(name="Calibri", size=10, color="1A73E8")
    alert_fill = PatternFill("solid", fgColor="FEF7E0")
    alert_font = Font(name="Calibri", size=10, bold=True, color="B06000")
    border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    safe_competitor = sanitize_excel_text(competitor_name, 80) or "Конкурент"
    safe_category = sanitize_excel_text(category_name, 120) or "Все категории"
    headers = [
        "Код нашего товара",
        "Наименование товара (Самбери)",
        "Распознано на ценнике (Конкурент)",
        "Цена закупки товара",
        "Цена продажи товара",
        "Цена на промо у товара",
        f"Текущая цена {safe_competitor}",
        f"Цена на промо {safe_competitor}",
        "Разница цен (руб)",
        "Price Index (PI %)",
        "Статус позиционирования",
        "Предупреждения",
    ]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"] = "САМБЕРИ: Отчёт по мониторингу ценников конкурентов"
    sheet["A1"].font = title_font
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet["A2"] = f"Категория: {safe_category} • Позиций: {len(items)}"
    sheet["A2"].font = meta_font

    header_row = 4
    for column_index, header in enumerate(headers, 1):
        cell = sheet.cell(header_row, column_index, sanitize_excel_text(header, 200))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[header_row].height = 32

    for row_index, item in enumerate(items, header_row + 1):
        warnings = [sanitize_excel_text(item.get("alert"), 500)] if item.get("alert") else []
        warnings.extend(
            sanitize_excel_text(warning, 500)
            for warning in item.get("data_quality_warnings", [])
            if warning
        )
        pi_percent = _safe_number(item.get("price_index_effective"))
        values = [
            sanitize_excel_text(item.get("matched_sku") or "-", 100),
            sanitize_excel_text(item.get("matched_name") or "-", 500),
            sanitize_excel_text(item.get("product_name"), 500),
            _safe_number(item.get("our_purchase_price")),
            _safe_number(item.get("our_sale_price")),
            _safe_number(item.get("our_promo_price")),
            _safe_number(_first_present(item, "comp_regular_price", "regular_price")),
            _safe_number(_first_present(item, "comp_promo_price", "promo_price")),
            _safe_number(item.get("effective_diff_rub")),
            pi_percent / 100.0 if pi_percent is not None else None,
            sanitize_excel_text(item.get("status"), 200),
            " | ".join(dict.fromkeys(filter(None, warnings))),
        ]

        for column_index, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column_index, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=column_index in {2, 3, 12})
            if column_index in {4, 5, 6, 7, 8, 9} and value is not None:
                cell.number_format = '#,##0.00 "₽"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif column_index == 10 and value is not None:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif column_index in {1, 11}:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if column_index == 11:
                status = str(value or "")
                if "Самбери дешевле" in status:
                    cell.fill, cell.font = green_fill, green_font
                elif "Конкурент дешевле" in status:
                    cell.fill, cell.font = red_fill, red_font
                elif "Паритет" in status:
                    cell.fill, cell.font = blue_fill, blue_font
            if column_index == 12 and value:
                cell.fill, cell.font = alert_fill, alert_font
        sheet.row_dimensions[row_index].height = 30

    last_row = max(header_row, header_row + len(items))
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A{header_row}:L{last_row}"
    sheet.print_title_rows = f"1:{header_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    widths = [18, 40, 40, 18, 18, 18, 18, 18, 18, 18, 26, 48]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    summary = summarize_price_index(items)
    summary_sheet = workbook.create_sheet("Сводка")
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet["A1"] = "Сводка мониторинга"
    summary_sheet["A1"].font = title_font
    summary_sheet["A2"] = f"Категория: {safe_category}"
    summary_sheet["A2"].font = meta_font
    summary_rows = [
        ("Всего ценников", summary["total_items"]),
        ("Успешно распознано", summary["successful_recognitions"]),
        ("Сопоставлено с каталогом", summary["matched_items"]),
        ("Сопоставимых цен", summary["comparable_items"]),
        ("Доля матчинга", summary["match_rate"] / 100.0),
        (
            "Средний Price Index",
            summary["avg_price_index"] / 100.0 if summary["avg_price_index"] is not None else None,
        ),
        (
            "Корзинный Price Index",
            summary["basket_price_index"] / 100.0
            if summary["basket_price_index"] is not None
            else None,
        ),
        ("Самбери дешевле", summary["samberi_cheaper_count"]),
        ("Конкурент дешевле", summary["competitor_cheaper_count"]),
        ("Паритет", summary["parity_count"]),
        ("Алерты демпинга", summary["dumping_alerts_count"]),
        ("Корзина Самбери", summary["total_our_basket"]),
        (f"Корзина {safe_competitor}", summary["total_comp_basket"]),
    ]
    for row_index, (label, value) in enumerate(summary_rows, 4):
        summary_sheet.cell(row_index, 1, sanitize_excel_text(label, 200))
        summary_sheet.cell(row_index, 2, value)
        summary_sheet.cell(row_index, 1).border = border
        summary_sheet.cell(row_index, 2).border = border
        if row_index in {8, 9, 10} and value is not None:
            summary_sheet.cell(row_index, 2).number_format = "0.0%"
        if row_index in {15, 16} and value is not None:
            summary_sheet.cell(row_index, 2).number_format = '#,##0.00 "₽"'
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 20

    buffer = io.BytesIO()
    workbook.save(buffer)
    excel_bytes = buffer.getvalue()
    if output_path:
        Path(output_path).write_bytes(excel_bytes)
    return excel_bytes
